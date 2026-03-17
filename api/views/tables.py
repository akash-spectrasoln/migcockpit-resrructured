"""
Table-related API views.
Handles file uploads, table creation, data manipulation, and table management.
"""
import io
import logging

from django.conf import settings
from django.db import connection
from django.http import HttpResponse
from django.utils import timezone
import pandas as pd
import psycopg2
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.serializers import TokenRefreshSerializer

from api.authentications import JWTCookieAuthentication
from api.connections.source_config import (
    convert_user_date_format_to_strftime,
    format_date_columns,
)
from api.models import Customer, User
from api.serializers import FileUploadSerializer

logger = logging.getLogger(__name__)

class FileUploadPreviewView(APIView):
    """
    API view for uploading files and previewing table structure without creating the table.
    This allows users to review and modify column types before table creation.
"""

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value

        str_value = str(value).strip()

        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)

                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return f"{float_val:.10f}".rstrip('0').rstrip('.')
                else:
                    # Default: convert to int (for backward compatibility with phone numbers)
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value

        return str_value

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()

        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)

        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')

        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)

        return df_cleaned

    def _detect_date_format(self, series):
        """
        Detect the consistent date format for a column by analyzing unambiguous dates.
        Returns the format string that should be used, or None if no format detected.
        """
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return None

        # Sample for performance
        sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values

        # Try each format and score based on successful conversions
        formats_to_try = [
            '%Y-%m-%d',                  # 2024-01-15
            '%m/%d/%Y',                  # 01/15/2024 (US format - month first)
            '%d/%m/%Y',                  # 15/01/2024 (European format - day first)
            '%m-%d-%Y',                  # 01-15-2024 (US format - month first)
            '%d-%m-%Y',                  # 15-01-2024 (European format - day first)
            '%Y/%m/%d',                  # 2024/01/15
            '%Y%m%d',                    # 20240115
            '%d.%m.%Y',                  # 15.01.2024
            '%Y-%m-%d %H:%M:%S',         # 2024-01-15 14:30:00
            '%m/%d/%Y %H:%M:%S',         # 01/15/2024 14:30:00
            '%d/%m/%Y %H:%M:%S',         # 15/01/2024 14:30:00
            '%m-%d-%Y %H:%M:%S',         # 01-15-2024 14:30:00
            '%d-%m-%Y %H:%M:%S',         # 15-01-2024 14:30:00
            '%Y-%m-%d %H:%M:%S.%',      # 2024-01-15 14:30:00.123456
            '%m-%d-%Y %H:%M:%S.%',      # 01-15-2024 14:30:00.123456
            '%d-%m-%Y %H:%M:%S.%',      # 15-01-2024 14:30:00.123456
            '%m/%d/%Y %H:%M:%S.%',      # 01/15/2024 14:30:00.123456
            '%d/%m/%Y %H:%M:%S.%',      # 15/01/2024 14:30:00.123456
        ]

        best_format = None
        best_score = 0

        for fmt in formats_to_try:
            try:
                converted = pd.to_datetime(sample, format=fmt, errors='coerce')
                success_count = converted.notna().sum()

                # Calculate success rate
                if success_count > best_score:
                    best_score = success_count
                    best_format = fmt
            except (ValueError, TypeError):
                continue

        # Return format if at least 60% success rate
        if best_score / len(sample) > 0.6:
            return best_format

        return None

    def _is_datetime_column(self, series):
        """
        Efficiently detect if a pandas Series contains datetime values.
        Uses format detection to ensure consistent parsing.
        """
        # Skip empty or all-null columns
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return False

        # If already datetime type, return True
        if pd.api.types.is_datetime64_any_dtype(series):
            return True

        # Try to detect the format
        detected_format = self._detect_date_format(series)
        if detected_format:
            return True

        # Fallback: Try with infer_datetime_format
        try:
            sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values
            converted = pd.to_datetime(sample, errors='coerce', infer_datetime_format=True)
            success_rate = converted.notna().sum() / len(sample)
            return success_rate > 0.6
        except (ValueError, TypeError):
            return False

    def _detect_datetime_type(self, series):
        """
        Determine if datetime column should be DATE or TIMESTAMP.
        Returns 'DATE' if all times are midnight (00:00:00), otherwise 'TIMESTAMP'.
        Uses consistent format detection.
        """
        try:
            # First, try to detect the consistent format
            detected_format = self._detect_date_format(series)

            if detected_format:
                # Use the detected format for consistent parsing
                datetime_series = pd.to_datetime(series, format=detected_format, errors='coerce')
            else:
                # Fallback to infer_datetime_format
                datetime_series = pd.to_datetime(series, errors='coerce', infer_datetime_format=True)

            non_null_datetimes = datetime_series.dropna()

            if len(non_null_datetimes) == 0:
                return 'TIMESTAMP'  # Default to TIMESTAMP if no valid dates

            # Sample for performance (check first 100 rows)
            sample = non_null_datetimes.head(100) if len(non_null_datetimes) > 100 else non_null_datetimes

            # Check if all times are at midnight (00:00:00)
            # This indicates date-only data
            has_time_component = any(
                dt.hour != 0 or dt.minute != 0 or dt.second != 0 or dt.microsecond != 0
                for dt in sample
            )

            # Return DATE if no time component, TIMESTAMP if has time
            return 'TIMESTAMP' if has_time_component else 'DATE'
        except Exception:
            return 'TIMESTAMP'  # Default to TIMESTAMP on error

    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)

        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            file_extension = uploaded_file.name.lower().split('.')[-1]

            # Get user and customer information
            try:
                user = request.user
                customer = user.cust_id
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            try:
                uploaded_file.seek(0)

                # Read file into pandas DataFrame
                # Don't use dtype=str to allow pandas to infer types (especially dates)
                if file_extension == 'csv':
                    df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
                elif file_extension in ['xls', 'xlsx']:
                    df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
                else:
                    return Response(
                        {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Remove empty rows from DataFrame
                df = self._remove_empty_rows(df)

                # Clean DataFrame
                df_clean = df.where(pd.notnull(df), None)

                # Check for reserved field names
                reserved_fields = ['__id', 'is_active']
                found_reserved_fields = []
                for col in df_clean.columns:
                    col_lower = col.lower().strip()
                    if col_lower in reserved_fields or col_lower == 'id':
                        found_reserved_fields.append(col)

                if found_reserved_fields:
                    return Response(
                        {"error": f"The following field names are reserved and cannot be used: {', '.join(found_reserved_fields)}. Please rename these columns in your file and try again."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Fix scientific notation for all columns that might contain it
                for col in df_clean.columns:
                    # Check if any value in this column contains scientific notation
                    has_scientific_notation = False
                    for val in df_clean[col].dropna():
                        if isinstance(val, str) and 'E' in val.upper():
                            has_scientific_notation = True
                            break

                    if has_scientific_notation:
                        # Convert scientific notation back to regular numbers
                        df_clean[col] = df_clean[col].apply(lambda x: self._fix_scientific_notation(x) if pd.notna(x) else x)

                # Define table name with filename__ format
                base_name = uploaded_file.name.rsplit('.', 1)[0]
                table_name = f"{base_name.lower()}__".replace(" ", "_").replace("-", "_")

                # Detect column types by checking all data in each column
                column_info = []
                datetime_columns = {}  # Store columns that are datetime and their detected format

                for col in df_clean.columns:
                    col_name = col.strip().replace(" ", "_").replace("-", "_").lower()
                    col_lower = col.lower()
                    # Get all non-null values for accurate type detection
                    all_values = df_clean[col].dropna()
                    # Still get sample values for display purposes
                    sample_values = all_values.head(10)

                    # Enhanced type detection using ALL values in the column
                    # First, calculate max length for VARCHAR sizing
                    max_length = df_clean[col].astype(str).str.len().max()
                    varchar_length = max_length + 50  # Add exactly 50 to the biggest data length

                    col_type = None  # Initialize as None to track if type was detected

                    # Priority 1: Check for date/timestamp columns using pandas
                    if self._is_datetime_column(df_clean[col]):
                        col_type = self._detect_datetime_type(df_clean[col])
                        # Store the detected format for this column
                        detected_format = self._detect_date_format(df_clean[col])
                        datetime_columns[col] = detected_format
                    # Priority 2: Check for phone numbers
                    elif any(keyword in col_lower for keyword in ['phone', 'mobile', 'tel', 'contact']):
                        col_type = 'VARCHAR(20)'
                    # Priority 3: Check for values with + sign (likely phone numbers with country code)
                    elif any('+' in str(val) for val in all_values if pd.notna(val)):
                        col_type = f'VARCHAR({varchar_length})'
                    # Priority 4: Check for integers
                    elif all(str(val).isdigit() for val in all_values if pd.notna(val) and str(val) != ''):
                        # Check if numbers are within INTEGER range (-2,147,483,648 to 2,147,483,647)
                        max_int = 2147483647
                        min_int = -2147483648
                        try:
                            numeric_values = [int(val) for val in all_values if pd.notna(val) and str(val) != '']
                            if all(min_int <= val <= max_int for val in numeric_values):
                                col_type = 'INTEGER'
                            else:
                                col_type = 'BIGINT'  # Use BIGINT for large integers
                        except (ValueError, OverflowError):
                            col_type = 'BIGINT'  # Use BIGINT if conversion fails
                    # Priority 5: Check for decimal numbers
                    elif all(str(val).replace('.', '').isdigit() and str(val).count('.') <= 1 for val in all_values if pd.notna(val) and str(val) != ''):
                        try:
                            decimal_values = [float(val) for val in all_values if pd.notna(val) and str(val) != '']
                            # Check if all values are whole numbers (no decimal part)
                            if all(val.is_integer() for val in decimal_values):
                                # Check if they fit in INTEGER range
                                int_values = [int(val) for val in decimal_values]
                                if all(min_int <= val <= max_int for val in int_values):
                                    col_type = 'INTEGER'
                                else:
                                    col_type = 'BIGINT'
                            else:
                                col_type = 'REAL'
                        except (ValueError, OverflowError):
                            col_type = 'REAL'
                    # Default: VARCHAR with proper sizing
                    else:
                        col_type = f'VARCHAR({varchar_length})'

                    column_info.append({
                        'original_name': col,
                        'column_name': col_name,
                        'postgresql_type': col_type,
                        'sample_values': [str(val) for val in sample_values.head(3).tolist()]
                    })

                # Convert datetime columns using the detected format for consistency
                for col, fmt in datetime_columns.items():
                    if fmt:
                        # Use the detected format to parse ALL values consistently
                        df_clean[col] = pd.to_datetime(df_clean[col], format=fmt, errors='coerce')
                    else:
                        # Fallback if format detection failed
                        df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce', infer_datetime_format=True)

                # Get sample data for preview (first 10 rows)
                sample_data = df_clean.head(10).to_dict('records')

                return Response({
                    "message": f"File uploaded successfully! Review the table structure below. Table will be created in customer database: {customer.cust_db}",
                    "table_info": {
                        "table_name": table_name,
                        "rows_count": len(df_clean),
                        "columns": column_info
                    },
                    "sample_data": sample_data,
                    "file_name": uploaded_file.name,
                    "file_extension": file_extension,
                    "customer_info": {
                        "cust_id": customer.cust_id,
                        "cust_db": customer.cust_db,
                        "customer_name": customer.name
                    }
                }, status=status.HTTP_200_OK)

            except Exception as e:
                return Response(
                    {"error": f"Error processing file: {e!s}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class WriteTableToDatabaseView(APIView):
    """
    API view for writing table structure and data to database without using session storage.
    This view receives the file, table structure, and creates the table with data in the user's customer database.
    """

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value

        str_value = str(value).strip()

        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)

                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return str(float_val)
                else:
                    # Default behavior: convert to int for phone numbers and other cases
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value

        return str_value

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()

        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)

        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')

        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)

        return df_cleaned

    def _detect_date_format(self, series):
        """
        Detect the consistent date format for a column by analyzing unambiguous dates.
        Returns the format string that should be used, or None if no format detected.
        """
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return None

        # Sample for performance
        sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values

        # Try each format and score based on successful conversions
        formats_to_try = [
            '%Y-%m-%d',                  # 2024-01-15
            '%m/%d/%Y',                  # 01/15/2024 (US format - month first)
            '%d/%m/%Y',                  # 15/01/2024 (European format - day first)
            '%m-%d-%Y',                  # 01-15-2024 (US format - month first)
            '%d-%m-%Y',                  # 15-01-2024 (European format - day first)
            '%Y/%m/%d',                  # 2024/01/15
            '%Y%m%d',                    # 20240115
            '%d.%m.%Y',                  # 15.01.2024
            '%Y-%m-%d %H:%M:%S',         # 2024-01-15 14:30:00
            '%m/%d/%Y %H:%M:%S',         # 01/15/2024 14:30:00
            '%d/%m/%Y %H:%M:%S',         # 15/01/2024 14:30:00
            '%m-%d-%Y %H:%M:%S',         # 01-15-2024 14:30:00
            '%d-%m-%Y %H:%M:%S',         # 15-01-2024 14:30:00
            '%Y-%m-%d %H:%M:%S.%',      # 2024-01-15 14:30:00.123456
            '%m-%d-%Y %H:%M:%S.%',      # 01-15-2024 14:30:00.123456
            '%d-%m-%Y %H:%M:%S.%',      # 15-01-2024 14:30:00.123456
            '%m/%d/%Y %H:%M:%S.%',      # 01/15/2024 14:30:00.123456
            '%d/%m/%Y %H:%M:%S.%',      # 15/01/2024 14:30:00.123456
        ]

        best_format = None
        best_score = 0

        for fmt in formats_to_try:
            try:
                converted = pd.to_datetime(sample, format=fmt, errors='coerce')
                success_count = converted.notna().sum()

                # Calculate success rate
                if success_count > best_score:
                    best_score = success_count
                    best_format = fmt
            except (ValueError, TypeError):
                continue

        # Return format if at least 60% success rate
        if best_score / len(sample) > 0.6:
            return best_format

        return None

    def _convert_datetime_value(self, value, column_type):
        """
        Convert datetime value to appropriate format for database insertion.
        Handles both DATE and TIMESTAMP types.
        """
        if pd.isna(value) or value == '' or value is None:
            return None

        try:
            # Convert to datetime using pandas
            dt = pd.to_datetime(value, errors='coerce', infer_datetime_format=True)

            if pd.isna(dt):
                return None

            # Convert pandas Timestamp to Python datetime
            if isinstance(dt, pd.Timestamp):
                dt = dt.to_pydatetime()

            # Return the datetime object (PostgreSQL will handle the conversion)
            return dt
        except Exception:
            return None

    def post(self, request):
        try:
            # Get the uploaded file
            if 'file' not in request.FILES:
                return Response(
                    {"error": "No file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            uploaded_file = request.FILES['file']
            table_name = request.data.get('table_name')
            columns_data = request.data.get('columns', [])
            scope = request.data.get('scope')
            schema = request.data.get('schema')

            # Parse columns if it's a JSON string
            if isinstance(columns_data, str):
                import json
                try:
                    columns = json.loads(columns_data)
                except json.JSONDecodeError:
                    return Response(
                        {"error": "Invalid columns data format."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                columns = columns_data

            if not table_name or not columns:
                return Response(
                    {"error": "Table name and columns are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Determine target schema based on scope
            if scope == 'global':
                target_schema = 'GENERAL'
            else:  # local scope
                target_schema = schema

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id  # user.cust_id is already the Customer object
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            file_extension = uploaded_file.name.lower().split('.')[-1]

            # Read file into pandas DataFrame
            # Don't use dtype=str to preserve datetime types
            uploaded_file.seek(0)
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Remove empty rows from DataFrame
            df = self._remove_empty_rows(df)

            # Clean DataFrame
            df_clean = df.where(pd.notnull(df), None)

            # Convert datetime columns using detected format for consistency
            # Check each column to see if it's a date/timestamp type
            for col in columns:
                original_name = col.get('original_name')
                col_type = col.get('postgresql_type', '').upper()

                # If this column is DATE or TIMESTAMP type, convert it consistently
                if col_type in ['DATE', 'TIMESTAMP', 'TIMESTAMP WITHOUT TIME ZONE', 'TIMESTAMP WITH TIME ZONE']:
                    if original_name in df_clean.columns:
                        # Detect format and convert
                        detected_format = self._detect_date_format(df_clean[original_name])
                        if detected_format:
                            df_clean[original_name] = pd.to_datetime(df_clean[original_name], format=detected_format, errors='coerce')
                        else:
                            df_clean[original_name] = pd.to_datetime(df_clean[original_name], errors='coerce', infer_datetime_format=True)

            # Build SQL for creating table
            column_definitions = []

            # Add system fields first
            column_definitions.append('__id SERIAL PRIMARY KEY')

            for col in columns:
                col_name = col.get('column_name')
                col_type = col.get('postgresql_type')
                if col_name and col_type:
                    column_definitions.append(f'"{col_name}" {col_type}')

            # Add is_active field at the end
            column_definitions.append('__active BOOLEAN DEFAULT TRUE')

            if not column_definitions:
                return Response(
                    {"error": "No valid column definitions provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(column_definitions)});'

            # Use transaction.on_commit to avoid transaction conflicts
            from django.db import transaction

            # Prepare data for table creation (outside transaction)
            # Filter out system fields (__id and is_active) from insert columns since they're auto-generated
            insert_columns = [col for col in columns if col.get('column_name') not in ['__id', 'is_active']]
            column_names = [col.get('column_name') for col in insert_columns]
            original_column_names = [col.get('original_name') for col in insert_columns]
            placeholders = ', '.join(['%s'] * len(column_names))
            f'INSERT INTO "{table_name}" ({", ".join(column_names)}) VALUES ({placeholders})'

            data_tuples = []
            for _, row in df_clean.iterrows():
                row_data = []
                for i, col_name in enumerate(column_names):
                    original_col_name = original_column_names[i]
                    col_type = insert_columns[i].get('postgresql_type', '') if i < len(insert_columns) else None

                    # Handle is_active field specially
                    if col_name == 'is_active':
                        row_data.append(True)  # Default to True for is_active
                    else:
                        # Try to get the value using the original column name
                        if original_col_name in df_clean.columns:
                            val = row[original_col_name]
                        else:
                            # If original column name not found, try the database column name
                            val = row[col_name] if col_name in df_clean.columns else None

                        if pd.isna(val):
                            row_data.append(None)
                        # Handle date/timestamp columns
                        elif col_type and col_type.upper() in ['DATE', 'TIMESTAMP', 'TIMESTAMP WITHOUT TIME ZONE', 'TIMESTAMP WITH TIME ZONE']:
                            row_data.append(self._convert_datetime_value(val, col_type))
                        elif isinstance(val, pd.Timestamp):
                            row_data.append(val.to_pydatetime())
                        else:
                            # Apply scientific notation fix for phone numbers and numeric columns before storing
                            if (original_col_name and any(keyword in original_col_name.lower() for keyword in ['phone', 'mobile', 'tel', 'contact'])) or \
                               (col_type and col_type.upper() in ['BIGINT', 'INTEGER', 'SMALLINT', 'NUMERIC', 'DECIMAL']):
                                val = self._fix_scientific_notation(val, col_type)
                            row_data.append(val)
                data_tuples.append(tuple(row_data))

            def create_table_in_customer_db():
                # Connect to customer's database instead of default database
                customer_db_config = {
                    'ENGINE': 'django.db.backends.postgresql',
                    'NAME': customer.cust_db,
                    'USER': settings.DATABASES['default']['USER'],
                    'PASSWORD': settings.DATABASES['default']['PASSWORD'],
                    'HOST': settings.DATABASES['default']['HOST'],
                    'PORT': settings.DATABASES['default']['PORT'],
                }

                # Create connection to customer database
                customer_connection = psycopg2.connect(
                    host=customer_db_config['HOST'],
                    port=customer_db_config['PORT'],
                    database=customer_db_config['NAME'],
                    user=customer_db_config['USER'],
                    password=customer_db_config['PASSWORD']
                )

                # Set autocommit after connection is established
                customer_connection.autocommit = True

                with customer_connection.cursor() as cursor:
                    # Set search path to target schema
                    cursor.execute(f'SET search_path TO "{target_schema}";')

                    # Drop table if exists (in target schema)
                    cursor.execute(f'DROP TABLE IF EXISTS "{target_schema}"."{table_name}";')

                    # Create table in target schema
                    create_table_sql_with_schema = f'CREATE TABLE IF NOT EXISTS "{target_schema}"."{table_name}" ({", ".join(column_definitions)});'
                    cursor.execute(create_table_sql_with_schema)

                    # Insert data (in target schema)
                    insert_sql_with_schema = f'INSERT INTO "{target_schema}"."{table_name}" ({", ".join(column_names)}) VALUES ({placeholders})'
                    cursor.executemany(insert_sql_with_schema, data_tuples)

                    # Create the activity_log table if it does not exist
                    create_log_table_sql = '''
                        CREATE TABLE IF NOT EXISTS "GENERAL"."activity_log" (
                            table_name VARCHAR(100) NOT NULL,
                            created_by VARCHAR(100),
                            created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            records_count INTEGER NOT NULL,
                            modified_by VARCHAR(100),
                            modified_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            description TEXT
                        );
                    '''
                    cursor.execute(create_log_table_sql)

                    # Insert a log entry for the newly created table
                    insert_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_log_sql,
                        (f'{target_schema}.{table_name}', user.email, len(data_tuples), None, f"Table '{table_name}' created in {target_schema} schema with {len(data_tuples)} records")
                    )

                customer_connection.close()

            # Execute table creation after transaction commits
            transaction.on_commit(create_table_in_customer_db)

            return Response({
                "message": f"Table '{table_name}' will be created in customer database '{customer.cust_db}' in the {target_schema} schema with {len(columns)} columns and {len(data_tuples)} rows.",
                "table_name": table_name,

                "columns": [col.get('column_name') for col in columns],
                "rows_inserted": len(data_tuples),
                "customer_info": {
                    "cust_id": customer.cust_id,
                    "cust_db": customer.cust_db,
                    "customer_name": customer.name
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error creating table: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ListUploadedTablesView(APIView):
    """
    API view to list tables ending with '__' from the user's customer database GENERAL schema.
    Requires JWT authentication.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, project_id):
        try:
            # Get the authenticated user from the request
            user = request.user

            # Check if user is associated with a customer
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            customer = user.cust_id  # user.cust_id is already the Customer object

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # # Set search path to GENERAL schema
                # cursor.execute('SET search_path TO "GENERAL";')

                # Query for tables ending with '__' in GENERAL schema (global)
                cursor.execute("""
                    SELECT table_name
                    FROM information_schema.tables
                    WHERE table_schema = 'GENERAL'
                    AND table_name LIKE '%__'             -- Ends with two underscores
                    AND RIGHT(table_name, 2) = '__'       -- Enforces that the two underscores are at the end
                    ORDER BY table_name;
                """)
                tables_query = cursor.fetchall()

                global_tables = []
                for row in tables_query:
                    global_tables.append({
                        'table_name': row[0],
                        'schema': 'GENERAL'
                    })

                # Query for schemas containing the project_id in their name (local)
                cursor.execute("""
                    SELECT schema_name
                    FROM information_schema.schemata
                    WHERE schema_name ILIKE %s
                    ORDER BY schema_name;
                """, ['%' + project_id + '%'])
                schemas_query = cursor.fetchall()
                schemas = [schema[0] for schema in schemas_query]

                # If there is exactly one schema, fetch table names ending with '__' from that schema
                local_tables = []
                if len(schemas) == 1:
                    schema_name = schemas[0] # schema name is FAD_M01
                    cursor.execute("""
                        SELECT table_name
                        FROM information_schema.tables
                        WHERE table_schema = %s
                        AND RIGHT(table_name, 2) = '__'
                        ORDER BY table_name;
                    """, [schema_name])
                    local_tables_query = cursor.fetchall()

                    for row in local_tables_query:
                        local_tables.append({
                            'table_name': row[0],
                            'schema': schema_name
                        })

            customer_connection.close()

            return Response({
                "message": f"Found {len(global_tables)} tables ending with '__' in GENERAL schema.",
                "global_tables": global_tables,
                "local_tables": local_tables,
                "schema": schema_name,
                "customer_info": {
                    "cust_id": customer.cust_id,
                    "cust_db": customer.cust_db,
                    "customer_name": customer.name
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error listing tables: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class GetTableDataView(APIView):
    """
    API view to retrieve records from a specific table in the user's customer database with pagination.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _build_where_clause(self, filters, columns):
        """
        Build WHERE clause and parameters from filters list.
        Filters format: [{"column": "name", "operator": "=", "value": "John"}, ...]
        Supported operators: =, !=, >, <, >=, <=, LIKE, ILIKE, IN, NOT IN
        """
        if not filters:
            return "", []

        where_conditions = []
        where_params = []
        column_names = [col[0] for col in columns]

        for filter_item in filters:
            column = filter_item.get('column', '')
            operator = filter_item.get('operator', '=')
            value = filter_item.get('value', '')

            # Validate column exists
            if column not in column_names:
                continue

            # Handle different operators
            if operator == '=':
                where_conditions.append(f'"{column}" = %s')
                where_params.append(value)
            elif operator == '!=':
                where_conditions.append(f'"{column}" != %s')
                where_params.append(value)
            elif operator == '>':
                where_conditions.append(f'"{column}" > %s')
                where_params.append(value)
            elif operator == '<':
                where_conditions.append(f'"{column}" < %s')
                where_params.append(value)
            elif operator == '>=':
                where_conditions.append(f'"{column}" >= %s')
                where_params.append(value)
            elif operator == '<=':
                where_conditions.append(f'"{column}" <= %s')
                where_params.append(value)
            elif operator == 'LIKE':
                where_conditions.append(f'"{column}" LIKE %s')
                where_params.append(f'%{value}%')
            elif operator == 'ILIKE':
                where_conditions.append(f'"{column}" ILIKE %s')
                where_params.append(f'%{value}%')
            elif operator == 'IN':
                # Handle both list and comma-separated string
                if isinstance(value, list) and value:
                    placeholders = ','.join(['%s'] * len(value))
                    where_conditions.append(f'"{column}" IN ({placeholders})')
                    where_params.extend(value)
                elif isinstance(value, str) and value:
                    # Split comma-separated string into list
                    values_list = [v.strip() for v in value.split(',') if v.strip()]
                    if values_list:
                        placeholders = ','.join(['%s'] * len(values_list))
                        where_conditions.append(f'"{column}" IN ({placeholders})')
                        where_params.extend(values_list)
            elif operator == 'NOT IN':
                # Handle both list and comma-separated string
                if isinstance(value, list) and value:
                    placeholders = ','.join(['%s'] * len(value))
                    where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                    where_params.extend(value)
                elif isinstance(value, str) and value:
                    # Split comma-separated string into list
                    values_list = [v.strip() for v in value.split(',') if v.strip()]
                    if values_list:
                        placeholders = ','.join(['%s'] * len(values_list))
                        where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                        where_params.extend(values_list)
            elif operator == 'IS NULL':
                where_conditions.append(f'"{column}" IS NULL')
            elif operator == 'IS NOT NULL':
                where_conditions.append(f'"{column}" IS NOT NULL')
            elif operator == 'MISMATCH':
                # Handle MISMATCH operator - value contains JSON string with mismatch pairs
                try:
                    import json
                    mismatch_data = json.loads(value) if isinstance(value, str) else value
                    if isinstance(mismatch_data, list) and len(mismatch_data) > 0:
                        # Get the first item to determine related column name
                        first_item = mismatch_data[0]
                        related_column = None
                        for key in first_item.keys():
                            if key != column:
                                related_column = key
                                break

                        if related_column and related_column in column_names:
                            # Build OR conditions for each mismatch pair
                            mismatch_conditions = []
                            for mismatch_pair in mismatch_data:
                                mismatch_conditions.append(
                                    f'("{column}" = %s AND "{related_column}" = %s)'
                                )
                                where_params.append(mismatch_pair.get(column))
                                where_params.append(mismatch_pair.get(related_column))

                            # Join with OR
                            where_conditions.append(f'({" OR ".join(mismatch_conditions)})')
                except (json.JSONDecodeError, AttributeError, KeyError):
                    # If JSON parsing fails, skip this filter
                    continue

        if where_conditions:
            return f' WHERE {" AND ".join(where_conditions)}', where_params
        return "", []

    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            request.data.get('project_id')
            schema = request.data.get('schema')
            page = int(request.data.get('page', 1))
            page_size = int(request.data.get('page_size', 100))
            sort_column = request.data.get('sort_column', '')
            sort_direction = request.data.get('sort_direction', 'asc')  # 'asc' or 'desc'
            filters = request.data.get('filters', [])  # List of filter objects

            if not table_name:
                return Response(
                    {"error": "Table name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate pagination parameters
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 1000:  # Limit page size to prevent abuse
                page_size = 100

            # Validate sort parameters
            if sort_direction not in ['asc', 'desc']:
                sort_direction = 'asc'

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id  # user.cust_id is already the Customer object
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:

                # First, get column information including is_nullable
                cursor.execute("""
                    SELECT column_name, data_type, is_nullable
                    FROM information_schema.columns
                    WHERE table_schema = %s
                    AND table_name = %s
                    ORDER BY ordinal_position;
                """, (schema, table_name))

                columns = cursor.fetchall()
                if not columns:
                    return Response(
                        {"error": f"Table '{table_name}' not found or has no columns."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Build WHERE clause for filtering
                where_clause, where_params = self._build_where_clause(filters, columns)

                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'

                # Get total count of records for pagination info (with filters)
                count_query = f'SELECT COUNT(*) FROM {table_reference}{where_clause};'
                cursor.execute(count_query, where_params)
                total_records = cursor.fetchone()[0]

                # Calculate pagination
                offset = (page - 1) * page_size
                total_pages = (total_records + page_size - 1) // page_size  # Ceiling division

                # Build ORDER BY clause for sorting
                order_by_clause = ""
                if sort_column and sort_column in [col[0] for col in columns]:
                    # Validate that the sort column exists in the table
                    order_by_clause = f' ORDER BY "{sort_column}" {sort_direction.upper()}'

                # Get paginated data from the table with filtering and sorting
                query = f'SELECT * FROM {table_reference}{where_clause}{order_by_clause} LIMIT %s OFFSET %s;'
                cursor.execute(query, [*where_params, page_size, offset])
                rows = cursor.fetchall()

                # Convert rows to list of dictionaries
                data = []
                column_names = [col[0] for col in columns]

                for row in rows:
                    row_dict = {}
                    for i, value in enumerate(row):
                        row_dict[column_names[i]] = value
                    data.append(row_dict)

                # Format date/timestamp columns based on user's preferred date format
                user_strftime_format = convert_user_date_format_to_strftime(user.date_format)
                data = format_date_columns(data, columns, user_strftime_format)

                # Check if table has primary key
                cursor.execute("""
                    SELECT column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                        ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_schema = %s
                        AND tc.table_name = %s
                        AND tc.constraint_type = 'PRIMARY KEY'
                    ORDER BY kcu.ordinal_position;
                """, (schema, table_name))

                pk_columns = cursor.fetchall()
                has_primary_key = len(pk_columns) > 0

            customer_connection.close()

            # Get user's preferred file format
            user_file_format = user.file_format

            return Response({
                "message": f"Retrieved {len(data)} records from table '{table_name}' (page {page} of {total_pages}).",
                "table_name": table_name,
                "columns": [{"name": col[0], "type": col[1], "is_nullable": col[2]} for col in columns],
                "data": data,
                "pagination": {
                    "current_page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "total_records": total_records,
                    "has_next": page < total_pages,
                    "has_previous": page > 1,
                    "showing_start": offset + 1 if total_records > 0 else 0,
                    "showing_end": min(offset + page_size, total_records)
                },
                "sorting": {
                    "sort_column": sort_column,
                    "sort_direction": sort_direction
                },
                "filters": filters,
                "has_primary_key": has_primary_key,
                "primary_key_columns": [col[0] for col in pk_columns] if has_primary_key else [],
                "customer_info": {
                    "cust_id": customer.cust_id,
                    "cust_db": customer.cust_db,
                    "customer_name": customer.name
                },
                "user_file_format": user_file_format
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error retrieving table data: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class GetDistinctValuesView(APIView):
    """
    API view to get distinct values for a specific column in a table.
    Used for populating filter dropdowns.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            column_name = request.data.get('column_name')
            schema = request.data.get('schema')
            limit = int(request.data.get('limit', 1000))  # Limit distinct values returned

            if not table_name or not column_name:
                return Response(
                    {"error": "Table name and column name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # Check if column exists
                cursor.execute("""
                    SELECT column_name, data_type
                    FROM information_schema.columns
                    WHERE table_schema = %s
                    AND table_name = %s
                    AND column_name = %s
                """, (schema, table_name, column_name))

                column_info = cursor.fetchone()
                if not column_info:
                    return Response(
                        {"error": f"Column '{column_name}' not found in table '{table_name}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"' if schema else f'"{table_name}"'

                # Get distinct values
                cursor.execute(f'SELECT DISTINCT "{column_name}" FROM {table_reference} WHERE "{column_name}" IS NOT NULL ORDER BY "{column_name}" LIMIT %s;', (limit,))
                distinct_values = [row[0] for row in cursor.fetchall()]

                # Convert to string for JSON serialization
                distinct_values = [str(value) if value is not None else '' for value in distinct_values]

            customer_connection.close()

            return Response({
                "column_name": column_name,
                "data_type": column_info[1],
                "distinct_values": distinct_values,
                "total_count": len(distinct_values)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error retrieving distinct values: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class PreviewTableDataView(APIView):
    """
    API view to preview data from uploaded file for table insertion.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value

        str_value = str(value).strip()

        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)

                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return f"{float_val:.10f}".rstrip('0').rstrip('.')
                else:
                    # Default: convert to int (for backward compatibility with phone numbers)
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value

        return str_value

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()

        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)

        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')

        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)

        df_cleaned = df_cleaned.drop_duplicates()

        return df_cleaned

    def post(self, request):
        try:
            # Get the uploaded file
            if 'file' not in request.FILES:
                return Response(
                    {"error": "No file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            uploaded_file = request.FILES['file']
            table_name = request.data.get('table_name')
            request.data.get('schema') # Added schema extraction

            if not table_name:
                return Response(
                    {"error": "Table name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information from authenticated user
            user = request.user
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            file_extension = uploaded_file.name.lower().split('.')[-1]

            # Read file into pandas DataFrame
            # Don't use dtype=str to preserve datetime types
            uploaded_file.seek(0)
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Remove empty rows from DataFrame
            df = self._remove_empty_rows(df)

            # Clean DataFrame
            df_clean = df.where(pd.notnull(df), None)

            # Fix scientific notation for all columns that might contain it
            for col in df_clean.columns:
                # Check if any value in this column contains scientific notation
                has_scientific_notation = False
                for val in df_clean[col].dropna():
                    if isinstance(val, str) and 'E' in val.upper():
                        has_scientific_notation = True
                        break

                if has_scientific_notation:
                    # Convert scientific notation back to regular numbers
                    df_clean[col] = df_clean[col].apply(lambda x: self._fix_scientific_notation(x) if pd.notna(x) else x)

            # Convert to list of dictionaries for JSON serialization
            sample_data = df_clean.head(10).to_dict('records')

            return Response({
                "message": f"File preview successful. Found {len(df_clean)} rows of data.",
                "sample_data": sample_data,
                "total_rows": len(df_clean),
                "columns": list(df_clean.columns),
                "table_name": table_name
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error previewing data: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class UploadTableDataView(APIView):
    """
    API view to upload data to an existing table in the customer's database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def _fix_scientific_notation(self, value, data_type=None):
        """Convert scientific notation back to regular number format for numeric fields."""
        if pd.isna(value) or value == '' or value is None:
            return value

        str_value = str(value).strip()

        # Check if it's in scientific notation (contains 'E' or 'e')
        if 'E' in str_value.upper():
            try:
                # Convert to float first
                float_val = float(str_value)

                # Based on data type, convert appropriately
                if data_type and 'int' in data_type.lower():
                    # For integer types (bigint, integer, smallint), convert to int
                    int_val = int(float_val)
                    return str(int_val)
                elif data_type and ('numeric' in data_type.lower() or 'decimal' in data_type.lower()):
                    # For numeric/decimal types, keep as float but format properly
                    return f"{float_val:.10f}".rstrip('0').rstrip('.')
                else:
                    # Default: convert to int (for backward compatibility with phone numbers)
                    int_val = int(float_val)
                    return str(int_val)
            except (ValueError, OverflowError):
                # If conversion fails, return original value
                return str_value

        return str_value

    def _remove_duplicates_from_new_data(self, new_df):
        """
        Remove duplicates from new data
        """
        return new_df.drop_duplicates()

    def _detect_date_format(self, series):
        """
        Detect the consistent date format for a column by analyzing unambiguous dates.
        Returns the format string that should be used, or None if no format detected.
        """
        non_null_values = series.dropna()
        if len(non_null_values) == 0:
            return None

        # Sample for performance
        sample = non_null_values.head(100) if len(non_null_values) > 100 else non_null_values

        # Try each format and score based on successful conversions
        formats_to_try = [
            '%Y-%m-%d',                  # 2024-01-15
            '%m/%d/%Y',                  # 01/15/2024 (US format - month first)
            '%d/%m/%Y',                  # 15/01/2024 (European format - day first)
            '%m-%d-%Y',                  # 01-15-2024 (US format - month first)
            '%d-%m-%Y',                  # 15-01-2024 (European format - day first)
            '%Y/%m/%d',                  # 2024/01/15
            '%Y%m%d',                    # 20240115
            '%d.%m.%Y',                  # 15.01.2024
            '%Y-%m-%d %H:%M:%S',         # 2024-01-15 14:30:00
            '%m/%d/%Y %H:%M:%S',         # 01/15/2024 14:30:00
            '%d/%m/%Y %H:%M:%S',         # 15/01/2024 14:30:00
            '%m-%d-%Y %H:%M:%S',         # 01-15-2024 14:30:00
            '%d-%m-%Y %H:%M:%S',         # 15-01-2024 14:30:00
            '%Y-%m-%d %H:%M:%S.%',      # 2024-01-15 14:30:00.123456
            '%m-%d-%Y %H:%M:%S.%',      # 01-15-2024 14:30:00.123456
            '%d-%m-%Y %H:%M:%S.%',      # 15-01-2024 14:30:00.123456
            '%m/%d/%Y %H:%M:%S.%',      # 01/15/2024 14:30:00.123456
            '%d/%m/%Y %H:%M:%S.%',      # 15/01/2024 14:30:00.123456
        ]

        best_format = None
        best_score = 0

        for fmt in formats_to_try:
            try:
                converted = pd.to_datetime(sample, format=fmt, errors='coerce')
                success_count = converted.notna().sum()

                # Calculate success rate
                if success_count > best_score:
                    best_score = success_count
                    best_format = fmt
            except (ValueError, TypeError):
                continue

        # Return format if at least 60% success rate
        if best_score / len(sample) > 0.6:
            return best_format

        return None

    def _convert_datetime_value(self, value, column_type):
        """
        Convert datetime value to appropriate format for database insertion.
        Handles both DATE and TIMESTAMP types.
        """
        if pd.isna(value) or value == '' or value is None:
            return None

        try:
            # Convert to datetime using pandas
            dt = pd.to_datetime(value, errors='coerce', infer_datetime_format=True)

            if pd.isna(dt):
                return None

            # Convert pandas Timestamp to Python datetime
            if isinstance(dt, pd.Timestamp):
                dt = dt.to_pydatetime()

            # Return the datetime object (PostgreSQL will handle the conversion)
            return dt
        except Exception:
            return None

    def _remove_empty_rows(self, df):
        """
        Remove completely empty rows and duplicate rows from DataFrame
        """
        # Remove rows where all values are null, empty string, or whitespace
        df_cleaned = df.copy()

        # Replace empty strings and whitespace-only strings with NaN
        df_cleaned = df_cleaned.replace(r'^\s*$', None, regex=True)

        # Remove rows where all values are null/NaN
        df_cleaned = df_cleaned.dropna(how='all')

        # Remove duplicate rows (keep first occurrence)
        df_cleaned = df_cleaned.drop_duplicates()

        # Reset index to maintain sequential row numbers
        df_cleaned = df_cleaned.reset_index(drop=True)
        return df_cleaned

    def get_column_info(self, cursor, table_name, schema):
        """
        Get comprehensive column information from the database schema.
        Returns a dictionary with column names as keys and their details as values.
        """
        cursor.execute("""
            SELECT column_name, data_type, character_maximum_length, numeric_precision,
                   numeric_scale, datetime_precision, is_nullable
            FROM information_schema.columns
            WHERE table_schema = %s
            AND table_name = %s
            AND column_name NOT IN ('__active', '__id')
            ORDER BY ordinal_position;
        """, (schema, table_name))

        column_info = {}
        for col_name, data_type, char_max_length, numeric_precision, numeric_scale, datetime_precision, is_nullable in cursor.fetchall():
            column_info[col_name] = {
                'data_type': data_type,
                'char_max_length': char_max_length,
                'numeric_precision': numeric_precision,
                'numeric_scale': numeric_scale,
                'datetime_precision': datetime_precision,
                'is_nullable': is_nullable == 'YES'
            }

        return column_info

    def post(self, request):
        try:
            # Get the uploaded file
            if 'file' not in request.FILES:
                return Response(
                    {"error": "No file provided."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            uploaded_file = request.FILES['file']
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')

            if not table_name:
                return Response(
                    {"error": "Table name is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            file_extension = uploaded_file.name.lower().split('.')[-1]

            # Read file into pandas DataFrame
            # Don't use dtype=str to preserve datetime types
            uploaded_file.seek(0)
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8', keep_default_na=True)
            elif file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, engine='openpyxl', keep_default_na=True)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xls/.xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Remove empty rows from DataFrame
            df = self._remove_empty_rows(df)

            # Clean DataFrame
            df_clean = df.where(pd.notnull(df), None)

            # Check for reserved field names in the uploaded file
            reserved_fields = ['__id', '__active']
            found_reserved_fields = []
            for col in df_clean.columns:
                col_lower = col.lower().strip()
                if col_lower in reserved_fields or col_lower == 'id':
                    found_reserved_fields.append(col)

            if found_reserved_fields:
                return Response(
                    {"error": f"The following field names are reserved and cannot be used: {', '.join(found_reserved_fields)}. Please rename these columns in your file and try again."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Note: Scientific notation fix will be applied later based on actual database column types

            # Get table structure to match columns and detect datetime columns
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            # Don't set autocommit = True to allow savepoints
            customer_connection.autocommit = False

            with customer_connection.cursor() as cursor:
                # Get table columns (excluding __active and __id as they will be added automatically)
                cursor.execute("""
                    SELECT column_name, data_type
                    FROM information_schema.columns
                    WHERE table_schema = %s
                    AND table_name = %s
                    AND column_name NOT IN ('__active', '__id')
                    ORDER BY ordinal_position;
                """, (schema, table_name))

                table_columns = cursor.fetchall()

                if not table_columns:
                    customer_connection.close()
                    return Response(
                        {"error": f"Table '{table_name}' not found in schema '{schema}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Get comprehensive column information
                column_info = self.get_column_info(cursor, table_name, schema)

                # Validate column names match between file and database table
                db_column_names = [col[0].lower() for col in table_columns]
                file_column_names = [col.lower().strip() for col in df_clean.columns]

                # Find missing columns in file (database columns not found in file)
                missing_in_file = []
                for db_col in db_column_names:
                    if db_col not in file_column_names:
                        missing_in_file.append(db_col)

                # Find extra columns in file (file columns not found in database)
                extra_in_file = []
                for file_col in file_column_names:
                    if file_col not in db_column_names:
                        extra_in_file.append(file_col)

                # If there are missing or extra columns, return detailed error
                if missing_in_file or extra_in_file:
                    error_message = "Column mismatch between uploaded file and database table:\n"

                    if missing_in_file:
                        error_message += f"Missing columns in file: {', '.join(missing_in_file)}\n"

                    if extra_in_file:
                        error_message += f"Extra columns in file (not in database): {', '.join(extra_in_file)}\n"

                    error_message += f"\nDatabase table columns: {', '.join(db_column_names)}\n"
                    error_message += f"File columns: {', '.join(file_column_names)}"

                    customer_connection.close()
                    return Response(
                        {"error": error_message},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Map file columns to table columns (case-insensitive)
                column_mapping = {}
                [col.lower() for col in df_clean.columns]

                for db_col, _db_type in table_columns:
                    # Try to find matching column in file
                    matching_col = None
                    for file_col in df_clean.columns:
                        if file_col.lower() == db_col.lower():
                            matching_col = file_col
                            break

                    if matching_col:
                        column_mapping[db_col] = matching_col

                if not column_mapping:
                    customer_connection.close()
                    return Response(
                        {"error": "No matching columns found between file and table."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Convert datetime columns using detected format for consistency
                for db_col, file_col in column_mapping.items():
                    db_col_info = column_info.get(db_col, {})
                    db_data_type = db_col_info.get('data_type', '').lower()

                    # If this column is date/timestamp type, convert it consistently
                    if db_data_type in ['date', 'timestamp', 'timestamp without time zone', 'timestamp with time zone']:
                        # Detect format and convert
                        detected_format = self._detect_date_format(df_clean[file_col])
                        if detected_format:
                            df_clean[file_col] = pd.to_datetime(df_clean[file_col], format=detected_format, errors='coerce')
                        else:
                            df_clean[file_col] = pd.to_datetime(df_clean[file_col], errors='coerce', infer_datetime_format=True)

                # Use the original data (no truncation needed)
                processed_df = df_clean

                # Prepare data for insertion
                insert_columns = list(column_mapping.keys())
                placeholders = ', '.join(['%s'] * (len(insert_columns) + 1))  # +1 for __active
                table_reference = f'"{schema}"."{table_name}"'
                insert_sql = f'INSERT INTO {table_reference} ({", ".join(insert_columns)}, __active) VALUES ({placeholders})'

                data_tuples = []
                for _, row in processed_df.iterrows():
                    row_data = []
                    for db_col in insert_columns:
                        file_col = column_mapping[db_col]
                        val = row[file_col]

                        if pd.isna(val):
                            row_data.append(None)
                        else:
                            # Get database column information
                            db_col_info = column_info.get(db_col, {})
                            db_data_type = db_col_info.get('data_type', '').lower()

                            # Handle date/timestamp columns
                            if db_data_type in ['date', 'timestamp', 'timestamp without time zone', 'timestamp with time zone']:
                                row_data.append(self._convert_datetime_value(val, db_data_type))
                            elif isinstance(val, pd.Timestamp):
                                row_data.append(val.to_pydatetime())
                            else:
                                # Apply scientific notation fix based on database column type
                                # Check if this is a numeric column that might have scientific notation
                                if db_data_type and any(numeric_type in db_data_type for numeric_type in ['bigint', 'integer', 'smallint', 'numeric', 'decimal', 'real', 'double']):
                                    val = self._fix_scientific_notation(val, db_data_type)

                                # Send raw string value - let PostgreSQL handle type validation
                                row_data.append(str(val))

                    # Add __active = True for all rows
                    row_data.append(True)
                    data_tuples.append(tuple(row_data))

                # Start transaction for all insertions
                cursor.execute("BEGIN TRANSACTION")

                # Use batch processing for better performance
                successful_inserts = 0
                failed_inserts = 0
                failed_records = []
                batch_size = 100  # Process in batches of 100 records

                # Adjust batch size based on total records for optimal performance
                total_records = len(data_tuples)
                if total_records > 1000:
                    batch_size = 200
                elif total_records > 5000:
                    batch_size = 500
                elif total_records > 10000:
                    batch_size = 1000

                # Process all records in batches using savepoints
                for batch_start in range(0, len(data_tuples), batch_size):
                    batch_end = min(batch_start + batch_size, len(data_tuples))
                    batch_data = data_tuples[batch_start:batch_end]

                    # Create a savepoint for this batch
                    savepoint_name = f"batch_{batch_start}"
                    cursor.execute(f"SAVEPOINT {savepoint_name}")

                    try:
                        # Try to insert the entire batch
                        cursor.executemany(insert_sql, batch_data)
                        successful_inserts += len(batch_data)
                        # Release the savepoint since batch was successful
                        cursor.execute(f"RELEASE SAVEPOINT {savepoint_name}")

                    except Exception:
                        # If batch fails, rollback to savepoint and process individually
                        cursor.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")

                        # Process each record in the failed batch individually
                        for i, data_tuple in enumerate(batch_data):
                            # Create a savepoint for each individual record
                            record_savepoint = f"record_{batch_start}_{i}"
                            cursor.execute(f"SAVEPOINT {record_savepoint}")

                            try:
                                cursor.execute(insert_sql, data_tuple)
                                successful_inserts += 1
                                # Release the savepoint since record was successful
                                cursor.execute(f"RELEASE SAVEPOINT {record_savepoint}")

                            except Exception as insert_error:
                                # Rollback to the record savepoint
                                cursor.execute(f"ROLLBACK TO SAVEPOINT {record_savepoint}")
                                failed_inserts += 1

                                # Get the row data for error reporting
                                actual_row_index = batch_start + i
                                row_data = processed_df.iloc[actual_row_index]

                                # Get the first column value for identification
                                first_column_name = processed_df.columns[0]
                                first_column_value = row_data[first_column_name] if first_column_name in row_data else 'N/A'

                                failed_record_info = {
                                    'row_number': actual_row_index + 2,  # +2 because first row is header, so actual data starts from row 2
                                    'first_column_value': str(first_column_value),
                                    'error': str(insert_error),
                                    'data': {col: row_data[col] for col in processed_df.columns}
                                }
                                failed_records.append(failed_record_info)

                # Check if any records failed - if so, rollback entire transaction
                if failed_inserts > 0:
                    cursor.execute("ROLLBACK")
                    customer_connection.close()

                    # Prepare error response with all failed records
                    error_details = []
                    for failed_record in failed_records:
                        error_details.append(f"Row {failed_record['row_number']} ({failed_record['first_column_value']}): {failed_record['error'].split('LINE 1:')[0].strip()}")

                    error_message = f"Transaction rolled back. {failed_inserts} records failed to insert. "
                    if len(failed_records) > 10:
                        error_message += f"First 10 errors: {'; '.join(error_details[:10])}"
                    else:
                        error_message += f"All errors: {'; '.join(error_details)}"

                    return Response(
                        {"error": error_message, "failed_records": failed_records},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                else:
                    # All records succeeded - commit the transaction
                    cursor.execute("COMMIT")

                    # Log the file upload activity
                    try:
                        insert_activity_log_sql = '''
                            INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                            VALUES (%s, %s, %s, %s, %s)
                        '''
                        cursor.execute(
                            insert_activity_log_sql,
                            (table_name, None, successful_inserts, user.email, f"File uploaded to table '{table_name}' with {successful_inserts} records")
                        )
                        customer_connection.commit()
                    except Exception as log_error:
                        # Log error but don't fail the main operation
                        print(f"Warning: Failed to log activity: {log_error}")

            customer_connection.close()

            # All records inserted successfully (transaction was committed)
            response_data = {
                "message": f"Successfully uploaded {successful_inserts} rows to table '{table_name}'."
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            error_message = str(e)

            # Check if it's a PostgreSQL data type validation error
            if "invalid input syntax" in error_message.lower() or "invalid value" in error_message.lower():
                # Extract column and value information from the error
                if "for type" in error_message.lower():
                    return Response(
                        {"error": f"Data type validation error: {error_message}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            return Response(
                {"error": f"Error uploading data: {error_message}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class CreateTableRecordView(APIView):
    """
    API view to create a new record in a specific table in the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            request.data.get('schema')
            record_data = request.data.get('record_data')

            if not table_name or not record_data:
                return Response(
                    {"error": "Table name and record data are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # Build INSERT query
                columns = list(record_data.keys())
                ['%s'] * len(columns)
                values = list(record_data.values())

                # Construct the table reference with schema

                [f'"{col}"' for col in columns]
                insert_query = """
                    INSERT INTO {table_reference} ({', '.join(quoted_columns)})
                    VALUES ({', '.join(placeholders)})
                """

                cursor.execute(insert_query, values)

                # Log the record creation activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, 1, user.email, f"Record saved to table '{table_name}'")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")

            customer_connection.close()

            return Response({
                "message": f"Record created successfully in table '{table_name}'.",
                "table_name": table_name,
                "created_fields": columns
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            error_message = str(e)

            # Check if it's a PostgreSQL data type validation error
            if "invalid input syntax" in error_message.lower() or "invalid value" in error_message.lower():
                # Extract column and value information from the error
                if "for type" in error_message.lower():
                    return Response(
                        {"error": f"Data type validation error: {error_message}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            return Response(
                {"error": f"Error creating record: {error_message}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class EditTableRecordView(APIView):
    """
    API view to edit an existing record in a specific table in the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def put(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            record_id = request.data.get('record_id')
            record_data = request.data.get('record_data')

            if not table_name or not record_id or not record_data:
                return Response(
                    {"error": "Table name, record ID, and record data are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # First, get the primary key column name
                cursor.execute("""
                    SELECT column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                    ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_name = %s
                    AND tc.constraint_type = 'PRIMARY KEY'
                    AND tc.table_schema = %s
                """, [table_name, schema])

                pk_result = cursor.fetchone()
                if not pk_result:
                    return Response(
                        {"error": f"No primary key found for table '{table_name}'."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                pk_column = pk_result[0]

                # Get table column information for validation
                cursor.execute("""
                    SELECT column_name, data_type, character_maximum_length, is_nullable, column_default
                    FROM information_schema.columns
                    WHERE table_name = %s AND table_schema = %s
                    ORDER BY ordinal_position
                """, [table_name, schema])

                column_info = {row[0]: {
                    'data_type': row[1],
                    'max_length': row[2],
                    'is_nullable': row[3] == 'YES',
                    'default': row[4]
                } for row in cursor.fetchall()}

                # Validate record data
                validation_errors = []
                columns = list(record_data.keys())

                # Check if primary key is included
                if pk_column not in columns:
                    return Response(
                        {"error": f"Primary key column '{pk_column}' must be included in the update data."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Validate each field
                for col_name, col_value in record_data.items():
                    if col_name == pk_column:
                        continue  # Skip primary key validation

                    if col_name not in column_info:
                        validation_errors.append(f"Column '{col_name}' does not exist in table '{table_name}'")
                        continue

                    col_info = column_info[col_name]

                    # Handle NULL/empty values
                    is_null_or_empty = (col_value is None or col_value == '' or str(col_value).strip() == '')

                    if is_null_or_empty:
                        # Check if column allows NULL values
                        if not col_info['is_nullable'] and col_info['default'] is None:
                            validation_errors.append(f"Column '{col_name}' cannot be NULL or empty")
                        # Convert empty strings to None for database
                        record_data[col_name] = None

                        continue

                    # Validate data type for non-null values
                    type_error = self._validate_data_type(col_name, col_value, col_info['data_type'])
                    if type_error:
                        validation_errors.append(type_error)

                    # Validate field length for character types
                    if col_info['max_length'] and col_info['data_type'] in ['character varying', 'character', 'text']:
                        if len(str(col_value)) > col_info['max_length']:
                            validation_errors.append(f"Column '{col_name}' value exceeds maximum length of {col_info['max_length']} characters")

                # Return validation errors if any
                if validation_errors:
                    return Response(
                        {"error": "Validation failed", "details": validation_errors},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Build UPDATE query
                set_clauses = []
                values = []
                for col in columns:
                    if col != pk_column:
                        set_clauses.append(f'"{col}" = %s')
                        values.append(record_data[col])

                values.append(record_id)  # Add the record ID for WHERE clause

                # Construct the table reference with schema

                update_query = """
                    UPDATE {table_reference}
                    SET {', '.join(set_clauses)}
                    WHERE "{pk_column}" = %s
                """

                cursor.execute(update_query, values)

                if cursor.rowcount == 0:
                    return Response(
                        {"error": f"No record found with ID '{record_id}' in table '{table_name}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Log the record update activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, 1, user.email, f"Record updated in table '{table_name}' (ID: {record_id})")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")

            customer_connection.close()

            return Response({
                "message": f"Record updated successfully in table '{table_name}'.",
                "table_name": table_name,
                "record_id": record_id,
                "updated_fields": [col for col in columns if col != pk_column]
            }, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"EditTableRecordView Error: {e!s}")
            print(f"Traceback: {error_details}")
            return Response(
                {"error": f"Error updating record: {e!s}", "debug": error_details},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _validate_data_type(self, column_name, value, data_type):
        """
        Validate that a value matches the expected data type for a column.
        Returns None if valid, or an error message if invalid.
        """
        try:
            # Skip validation for None values (handled separately)
            if value is None:
                return None

            data_type_lower = data_type.lower()

            # Integer types
            if data_type_lower in ['integer', 'bigint', 'smallint', 'serial', 'bigserial']:
                try:
                    # Try to convert to int, handle both string and numeric inputs
                    int(float(str(value)))  # Convert to float first to handle "123.0" -> 123
                except (ValueError, TypeError, OverflowError):
                    return f"Column '{column_name}' must be an integer"

            # Numeric/Decimal types
            elif data_type_lower in ['numeric', 'decimal', 'real', 'double precision', 'float']:
                try:
                    float(str(value))
                except (ValueError, TypeError, OverflowError):
                    return f"Column '{column_name}' must be a number"
            # Boolean type
            elif data_type_lower == 'boolean':
                str_value = str(value).lower().strip()
                if str_value not in ['true', 'false', '1', '0', 'yes', 'no', 'on', 'of', 't', '']:
                    return f"Column '{column_name}' must be a boolean value (true/false)"

            # Date types
            elif data_type_lower == 'date':
                try:
                    from datetime import datetime
                    str_value = str(value).strip()
                    # Try different date formats
                    formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']
                    parsed = False
                    for fmt in formats:
                        try:
                            datetime.strptime(str_value, fmt)
                            parsed = True
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        return f"Column '{column_name}' must be a valid date (YYYY-MM-DD format)"
                except Exception:
                    return f"Column '{column_name}' must be a valid date"

            # Timestamp types
            elif data_type_lower in ['timestamp', 'timestamp without time zone', 'timestamp with time zone']:
                try:
                    from datetime import datetime
                    str_value = str(value).strip()
                    # Try different timestamp formats
                    formats = [
                        '%Y-%m-%d %H:%M:%S',
                        '%Y-%m-%dT%H:%M:%S',
                        '%Y-%m-%d %H:%M:%S.%',
                        '%Y-%m-%dT%H:%M:%S.%',
                        '%m/%d/%Y %H:%M:%S',
                        '%d/%m/%Y %H:%M:%S'
                    ]
                    parsed = False
                    for fmt in formats:
                        try:
                            datetime.strptime(str_value, fmt)
                            parsed = True
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        return f"Column '{column_name}' must be a valid timestamp"
                except Exception:
                    return f"Column '{column_name}' must be a valid timestamp"

            # Time types
            elif data_type_lower == 'time':
                try:
                    from datetime import datetime
                    str_value = str(value).strip()
                    formats = ['%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p']
                    parsed = False
                    for fmt in formats:
                        try:
                            datetime.strptime(str_value, fmt)
                            parsed = True
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        return f"Column '{column_name}' must be a valid time (HH:MM:SS format)"
                except Exception:
                    return f"Column '{column_name}' must be a valid time"

            # UUID type
            elif data_type_lower == 'uuid':
                try:
                    import uuid
                    uuid.UUID(str(value).strip())
                except (ValueError, TypeError):
                    return f"Column '{column_name}' must be a valid UUID"

            # JSON type
            elif data_type_lower in ['json', 'jsonb']:
                try:
                    import json
                    json.loads(str(value).strip())
                except (ValueError, TypeError):
                    return f"Column '{column_name}' must be valid JSON"

            # Text types (character varying, text, etc.) - no specific validation needed
            # as they can accept any string

            return None  # No validation error

        except Exception as e:
            return f"Error validating column '{column_name}': {e!s}"

class DeleteTableRecordView(APIView):
    """
    API view to delete a record from a specific table in the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            record_id = request.data.get('record_id')

            if not table_name or not record_id:
                return Response(
                    {"error": "Table name and record ID are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # First, get the primary key column name
                cursor.execute("""
                    SELECT column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                    ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_name = %s
                    AND tc.constraint_type = 'PRIMARY KEY'
                    AND tc.table_schema = %s
                """, [table_name, schema])

                pk_result = cursor.fetchone()
                if not pk_result:
                    return Response(
                        {"error": f"No primary key found for table '{table_name}'."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                pk_result[0]

                # Construct the table reference with schema

                # Build DELETE query
                delete_query = """
                    DELETE FROM {table_reference}
                    WHERE "{pk_column}" = %s
                """

                cursor.execute(delete_query, [record_id])

                if cursor.rowcount == 0:
                    return Response(
                        {"error": f"No record found with ID '{record_id}' in table '{table_name}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Log the record deletion activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, 1, user.email, f"Record deleted from table '{table_name}' (ID: {record_id})")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")

            customer_connection.close()

            return Response({
                "message": f"Record deleted successfully from table '{table_name}'.",
                "table_name": table_name,
                "record_id": record_id
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error deleting record: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

#updating user added table structure
class UpdateTableStructureView(APIView):
    """
    API view to update table structure by modifying column data types.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            schema = request.data.get('schema')
            table_name = request.data.get('table_name')
            column_changes = request.data.get('column_changes', {})
            column_renames = request.data.get('column_renames', {})
            columns_to_add = request.data.get('columns_to_add', [])
            columns_to_delete = request.data.get('columns_to_delete', [])

            if not schema or not table_name:
                return Response(
                    {"error": "Schema and table name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Check if at least one operation is requested
            if not column_changes and not column_renames and not columns_to_add and not columns_to_delete:
                return Response(
                    {"error": "At least one operation (column changes, renames, additions, or deletions) is required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # Check if table exists
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables
                        WHERE table_schema = %s
                        AND table_name = %s
                    )
                """, [schema, table_name])

                if not cursor.fetchone()[0]:
                    return Response(
                        {"error": f"Table '{table_name}' does not exist."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Get current column information
                cursor.execute("""
                    SELECT column_name, data_type, character_maximum_length
                    FROM information_schema.columns
                    WHERE table_schema = %s
                    AND table_name = %s
                    ORDER BY ordinal_position
                """, [schema, table_name])

                current_columns = cursor.fetchall()
                column_info = {col[0]: {'type': col[1], 'length': col[2]} for col in current_columns}

                # Process column operations
                updated_columns = []
                added_columns = []
                deleted_columns = []
                renamed_columns = []
                data_loss_warnings = []
                final_table_name = table_name

                # 1. Delete columns first (to avoid dependency issues)
                for column_name in columns_to_delete:
                    # Prevent deletion of system columns
                    if column_name in ['__id', '__active']:
                        data_loss_warnings.append(f"Cannot delete system column '{column_name}'.")
                        continue

                    if column_name not in column_info:
                        data_loss_warnings.append(f"Column '{column_name}' does not exist in table '{table_name}'.")
                        continue

                    # Check if column is part of primary key
                    cursor.execute("""
                        SELECT column_name
                        FROM information_schema.table_constraints tc
                        JOIN information_schema.key_column_usage kcu
                        ON tc.constraint_name = kcu.constraint_name
                        WHERE tc.table_name = %s
                        AND tc.constraint_type = 'PRIMARY KEY'
                        AND tc.table_schema = %s
                        AND kcu.column_name = %s
                    """, [table_name, schema, column_name])

                    if cursor.fetchone():
                        data_loss_warnings.append(f"Cannot delete column '{column_name}' as it is part of the primary key.")
                        continue

                    try:
                        drop_query = f'ALTER TABLE "{schema}"."{final_table_name}" DROP COLUMN "{column_name}"'
                        cursor.execute(drop_query)
                        deleted_columns.append(column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to delete column '{column_name}': {e!s}")

                # 2. Rename existing columns
                for old_column_name, new_column_name in column_renames.items():
                    if old_column_name not in column_info:
                        data_loss_warnings.append(f"Column '{old_column_name}' does not exist in table '{final_table_name}'.")
                        continue

                    # Clean and validate new column name
                    new_column_name = new_column_name.strip().replace(' ', '_')
                    if not new_column_name.replace('_', '').isalnum():
                        data_loss_warnings.append(f"Invalid column name '{new_column_name}'. Only alphanumeric characters and underscores are allowed.")
                        continue

                    # Check if new column name already exists
                    if new_column_name in column_info:
                        data_loss_warnings.append(f"Column '{new_column_name}' already exists in table '{final_table_name}'.")
                        continue

                    try:
                        rename_query = f'ALTER TABLE "{schema}"."{final_table_name}" RENAME COLUMN "{old_column_name}" TO "{new_column_name}"'
                        cursor.execute(rename_query)
                        renamed_columns.append(f"{old_column_name} -> {new_column_name}")

                        # Update column_info for subsequent operations
                        column_info[new_column_name] = column_info.pop(old_column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to rename column '{old_column_name}' to '{new_column_name}': {e!s}")

                # 3. Add new columns
                for column_data in columns_to_add:
                    column_name = column_data.get('name', '').strip().replace(' ', '_')
                    column_type = column_data.get('type', 'VARCHAR(255)').strip()

                    if not column_name:
                        data_loss_warnings.append("Column name cannot be empty.")
                        continue

                    # Validate column name (basic validation)
                    if not column_name.replace('_', '').isalnum():
                        data_loss_warnings.append(f"Invalid column name '{column_name}'. Only alphanumeric characters and underscores are allowed.")
                        continue

                    # Check if column already exists
                    if column_name in column_info:
                        data_loss_warnings.append(f"Column '{column_name}' already exists in table '{final_table_name}'.")
                        continue

                    try:
                        add_query = f'ALTER TABLE "{schema}"."{final_table_name}" ADD COLUMN "{column_name}" {column_type}'
                        cursor.execute(add_query)
                        added_columns.append(column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to add column '{column_name}': {e!s}")

                # 4. Update existing columns
                for column_name, new_type in column_changes.items():
                    if column_name not in column_info:
                        data_loss_warnings.append(f"Column '{column_name}' does not exist in table '{final_table_name}'.")
                        continue

                    # Skip if the type hasn't changed
                    current_type = column_info[column_name]['type']
                    if current_type.upper() == new_type.upper():
                        continue

                    # Check for potentially data-losing conversions
                    current_type_upper = current_type.upper()
                    new_type_upper = new_type.upper()

                    # Check for decimal to integer conversions
                    if (current_type_upper in ['REAL', 'DOUBLE PRECISION', 'NUMERIC', 'DECIMAL'] and
                        new_type_upper in ['INTEGER', 'BIGINT', 'SMALLINT', 'SERIAL', 'BIGSERIAL']):
                        data_loss_warnings.append(f"Converting '{column_name}' from {current_type} to {new_type} will truncate decimal values and cause data loss.")
                        continue

                    # Check for text to numeric conversions
                    if (current_type_upper in ['TEXT', 'VARCHAR', 'CHARACTER VARYING'] and
                        new_type_upper in ['INTEGER', 'BIGINT', 'SMALLINT', 'REAL', 'NUMERIC', 'DECIMAL']):
                        data_loss_warnings.append(f"Converting '{column_name}' from {current_type} to {new_type} may fail if text contains non-numeric values.")
                        continue

                    # Check for numeric to text conversions (usually safe but warn)
                    if (current_type_upper in ['INTEGER', 'BIGINT', 'SMALLINT', 'REAL', 'NUMERIC', 'DECIMAL'] and
                        new_type_upper in ['TEXT', 'VARCHAR', 'CHARACTER VARYING']):
                        data_loss_warnings.append(f"Converting '{column_name}' from {current_type} to {new_type} will convert numbers to text format.")
                        continue

                    # Build ALTER COLUMN statement
                    alter_query = f'ALTER TABLE "{schema}"."{final_table_name}" ALTER COLUMN "{column_name}" TYPE {new_type}'

                    try:
                        cursor.execute(alter_query)
                        updated_columns.append(column_name)
                    except Exception as e:
                        data_loss_warnings.append(f"Failed to update column '{column_name}' to type '{new_type}': {e!s}")

            customer_connection.close()

            # Check if any operations were successful
            total_operations = len(updated_columns) + len(added_columns) + len(deleted_columns) + len(renamed_columns)

            if total_operations == 0 and not data_loss_warnings:
                return Response({
                    "message": "No changes were made to the table structure.",
                    "table_name": final_table_name
                }, status=status.HTTP_200_OK)

            response_data = {
                "table_name": final_table_name,
                "updated_columns": updated_columns,
                "added_columns": added_columns,
                "deleted_columns": deleted_columns,
                "renamed_columns": renamed_columns
            }

            if data_loss_warnings:
                response_data["warnings"] = data_loss_warnings
                if total_operations > 0:
                    operations = []
                    if renamed_columns:
                        operations.append(f"{len(renamed_columns)} column(s) renamed")
                    if updated_columns:
                        operations.append(f"{len(updated_columns)} column(s) modified")
                    if added_columns:
                        operations.append(f"{len(added_columns)} column(s) added")
                    if deleted_columns:
                        operations.append(f"{len(deleted_columns)} column(s) deleted")

                    response_data["message"] = f"Table structure updated successfully. {', '.join(operations)}. Some operations were skipped due to errors or potential data loss."
                else:
                    response_data["message"] = "No changes were made due to errors or potential data loss in the requested operations."
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            else:
                operations = []
                if renamed_columns:
                    operations.append(f"{len(renamed_columns)} column(s) renamed")
                if updated_columns:
                    operations.append(f"{len(updated_columns)} column(s) modified")
                if added_columns:
                    operations.append(f"{len(added_columns)} column(s) added")
                if deleted_columns:
                    operations.append(f"{len(deleted_columns)} column(s) deleted")

                response_data["message"] = f"Table structure updated successfully. {', '.join(operations)}."

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error updating table structure: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

#deleting user added tables
class DeleteTableView(APIView):
    """
    API view to delete a table from the user's customer database.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def delete(self, request):
        try:
            schema = request.data.get('schema')
            table_name = request.data.get('table_name')

            if not schema or not table_name:
                return Response(
                    {"error": "Schema and table name are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id
            except User.DoesNotExist:
                return Response(
                    {"error": "User not found."},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:
                return Response(
                    {"error": "Customer not found for the selected user."},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer's database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # Check if table exists
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables
                        WHERE table_schema = %s
                        AND table_name = %s
                    )
                """, [schema, table_name])

                if not cursor.fetchone()[0]:
                    return Response(
                        {"error": f"Table '{table_name}' does not exist."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Get record count before deletion
                cursor.execute(f'SELECT COUNT(*) FROM "{schema}"."{table_name}";')
                record_count = cursor.fetchone()[0]

                # Drop the table
                drop_query = f'DROP TABLE "{schema}"."{table_name}";'
                cursor.execute(drop_query)

                # Log the table deletion activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "{schema}"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, record_count, user.email, f"Table '{table_name}' deleted with {record_count} records")
                    )
                except Exception as log_error:
                    # Log error but don't fail the main operation
                    print(f"Warning: Failed to log activity: {log_error}")

            customer_connection.close()

            return Response({
                "message": f"Table '{table_name}' deleted successfully. {record_count} records were removed.",
                "table_name": table_name,
                "deleted_records": record_count
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error deleting table: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class CreateTableWithoutRecordsView(APIView):
    """
    API view to create a directory for a table in the user's customer database.
    """

    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self,request):

        table_name = request.data.get('table_name')
        columns = request.data.get('column_names')
        scope = request.data.get('scope')
        schema = request.data.get('schema')

        if not table_name or not columns or not schema:
            return Response(
                {"error": "Table name and column names and schema are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        #validate table name format
        clean_table_name = table_name.strip().lower().replace(" ",'_')
        if not clean_table_name.replace("_","").isalnum():
            return Response(
                {"error": "Table name must contain only letters, numbers, and underscores."},
            )

        if not clean_table_name[0].isalpha() or clean_table_name[0] == "_":
            return Response(
                {"error": "Table name must start with a letter or underscore."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Add __ suffix to table name
        clean_table_name = clean_table_name + "__"

        column_names = []
        for col in columns:
            col_name = col.get("column_name")
            if col_name.lower() in ['__id', 'is_active']: # reserved fields
                return Response(
                    {"error": f"Column name {col_name} is reserved."},
                )
            data_type = col.get('data_type', '')

            if not col_name or not data_type:
                return Response(
                    {"error": "Column name and data type are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if col_name.lower() in column_names:
                return Response(
                    {"error": f"duplicate column name: {col_name}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            column_names.append(col_name.lower())

        try:
            user = request.user
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer = user.cust_id  # already a Customer instance
        except User.DoesNotExist:
            return Response(
                {"error": "User not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Connect to customer's database
        customer_connection = psycopg2.connect(
            host=settings.DATABASES['default']['HOST'],
            port=settings.DATABASES['default']['PORT'],
            database=customer.cust_db,
            user=settings.DATABASES['default']['USER'],
            password=settings.DATABASES['default']['PASSWORD']
        )
        customer_connection.autocommit = True

        # Determine target schema based on scope
        target_schema = 'GENERAL' if scope == 'global' else schema

        with customer_connection.cursor() as cursor:
            # Set search path to target schema
            cursor.execute(f'SET search_path TO "{target_schema}";')

            # Check if table exists
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_schema = %s
                    AND table_name = %s
                )
            """, [target_schema, clean_table_name])

            if cursor.fetchone()[0]:
                return Response(
                    {"error": f"Table '{table_name}' already exists."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            #create the column definition - start with auto-increment ID as primary key
            column_definitions = ['__id SERIAL PRIMARY KEY']
            for col in columns:
                col_name = col.get("column_name")
                data_type = col.get('data_type', '')
                is_nullable = col.get('is_nullable', True)
                column_definitions.append(f'"{col_name}" {data_type} {("NOT NULL" if not is_nullable else "")}')

            create_table_sql = f'CREATE TABLE IF NOT EXISTS "{target_schema}"."{clean_table_name}" ({", ".join(column_definitions)});'
            cursor.execute(create_table_sql)

            #add is_active field to the table
            cursor.execute('''
                ALTER TABLE "{target_schema}"."{clean_table_name}" ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE;
            ''')

            #create the activity_log table if doesnt exist
            create_log_table_sql = '''
                CREATE TABLE IF NOT EXISTS "{target_schema}"."activity_log" (
                    table_name VARCHAR(100) NOT NULL,
                    created_by VARCHAR(100),
                    created_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    records_count INTEGER NOT NULL,
                    modified_by VARCHAR(100),
                    modified_on TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    description TEXT
                );
            '''
            cursor.execute(create_log_table_sql)

            #insert the log entry
            insert_log_sql = '''
                INSERT INTO "{target_schema}"."activity_log" (table_name, created_by, records_count, modified_by, description)
                VALUES (%s, %s, %s, %s, %s)
            '''
            cursor.execute(
                insert_log_sql,
                (clean_table_name, user.email, 0, None, f"Table '{clean_table_name}' created with {len(columns) + 1} columns (including auto-increment ID)")
            )

        customer_connection.close()

        return Response(
            {"message": f"Table '{clean_table_name}' created successfully in schema '{target_schema}' with {len(columns) + 1} columns (including auto-increment ID).",
            "schema":schema},
            status=status.HTTP_200_OK
        )

class ImportDataFromHanaView(APIView):
    def post(self, request):
        user_id = request.data.get('user_id')
        print(f"User ID received: {user_id}")

        if not user_id:
            return Response(
                {"message": "User ID is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(
            {"message": f"Import data from HANA for user ID: {user_id}"},
            status=status.HTTP_200_OK
        )

class DownloadTableDataView(APIView):
    """
    API endpoint to download table data in multiple formats (CSV, Excel, TSV) with filtering support
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            table_name = request.data.get('table_name')
            schema = request.data.get('schema')
            filters = request.data.get('filters', [])
            sort_column = request.data.get('sort_column', '')
            sort_direction = request.data.get('sort_direction', 'asc')
            file_format = request.data.get('format', 'csv').lower()  # Default to CSV
            selected_columns = request.data.get('selected_columns', None)  # Get selected columns

            if not table_name:
                return Response(
                    {"error": "table_name is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate file format
            if file_format not in ['csv', 'excel', 'xlsx', 'tsv']:
                return Response(
                    {"error": "Invalid format. Supported formats: csv, excel, xlsx, tsv"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get user and customer information
            try:
                user = request.user
                if not user.cust_id:
                    return Response(
                        {"error": "User is not associated with any customer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                customer = user.cust_id

                # Update user's file format preference if format is provided
                if file_format in ['csv', 'excel', 'xlsx', 'tsv']:
                    # Normalize excel/xlsx to 'excel'
                    normalized_format = 'excel' if file_format in ['excel', 'xlsx'] else file_format
                    # Use raw SQL to update file_format since user.pk is email (string), not integer id
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute("""
                                UPDATE "GENERAL"."user"
                                SET file_format = %s
                                WHERE email = %s
                            """, [normalized_format, user.email])
                    except Exception as save_error:
                        print(f"Warning: Could not update file_format: {save_error}")
                        # Continue anyway - this is not critical

            except User.DoesNotExist:
                return Response(
                    {"error": "User not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
            except Customer.DoesNotExist:

                return Response(
                    {"error": "Customer not found"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Connect to customer database
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=customer.cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # Get table columns (excluding __active and __id as they will be added automatically)
                cursor.execute("""
                    SELECT column_name, data_type
                    FROM information_schema.columns
                    WHERE table_schema = %s
                    AND table_name = %s
                    AND column_name NOT IN ('__active', '__id')
                    ORDER BY ordinal_position;
                """, (schema, table_name))

                table_columns = cursor.fetchall()

                if not table_columns:
                    customer_connection.close()
                    return Response(
                        {"error": f"Table '{table_name}' not found in schema '{schema}'."},
                        status=status.HTTP_404_NOT_FOUND
                    )

                # Filter columns based on selected_columns if provided
                if selected_columns and isinstance(selected_columns, list) and len(selected_columns) > 0:
                    # Create a dictionary for quick lookup
                    columns_dict = {col[0]: col for col in table_columns}

                    # Filter and reorder table_columns to match selected_columns order
                    filtered_columns = []
                    for col_name in selected_columns:
                        if col_name in columns_dict:
                            filtered_columns.append(columns_dict[col_name])

                    table_columns = filtered_columns

                    # If no columns match after filtering, return error
                    if not table_columns:
                        customer_connection.close()
                        return Response(
                            {"error": "None of the selected columns exist in the table."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                # Build WHERE clause for filters
                where_conditions = []
                filter_params = []

                for filter_item in filters:
                    column = filter_item.get('column')
                    operator = filter_item.get('operator')
                    value = filter_item.get('value')

                    if column and operator and value is not None:
                        if operator == 'ILIKE':
                            where_conditions.append(f'"{column}" ILIKE %s')
                            filter_params.append(f'%{value}%')
                        elif operator == 'IN':
                            # Handle both list and comma-separated string
                            if isinstance(value, list) and value:
                                placeholders = ','.join(['%s'] * len(value))
                                where_conditions.append(f'"{column}" IN ({placeholders})')
                                filter_params.extend(value)
                            elif isinstance(value, str) and value:
                                values_list = [v.strip() for v in value.split(',') if v.strip()]
                                if values_list:
                                    placeholders = ','.join(['%s'] * len(values_list))
                                    where_conditions.append(f'"{column}" IN ({placeholders})')
                                    filter_params.extend(values_list)
                        elif operator == 'NOT IN':
                            # Handle both list and comma-separated string
                            if isinstance(value, list) and value:
                                placeholders = ','.join(['%s'] * len(value))
                                where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                                filter_params.extend(value)
                            elif isinstance(value, str) and value:
                                values_list = [v.strip() for v in value.split(',') if v.strip()]
                                if values_list:
                                    placeholders = ','.join(['%s'] * len(values_list))
                                    where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                                    filter_params.extend(values_list)
                        elif operator == 'IS NULL':
                            where_conditions.append(f'"{column}" IS NULL')
                        elif operator == 'IS NOT NULL':
                            where_conditions.append(f'"{column}" IS NOT NULL')
                        elif operator == 'MISMATCH':
                            # Handle MISMATCH operator - value contains JSON string with mismatch pairs
                            try:
                                import json
                                mismatch_data = json.loads(value) if isinstance(value, str) else value
                                if isinstance(mismatch_data, list) and len(mismatch_data) > 0:
                                    # Get the first item to determine related column name
                                    first_item = mismatch_data[0]
                                    related_column = None
                                    for key in first_item.keys():
                                        if key != column:
                                            related_column = key
                                            break

                                    if related_column:
                                        # Build OR conditions for each mismatch pair
                                        mismatch_conditions = []
                                        for mismatch_pair in mismatch_data:
                                            mismatch_conditions.append(
                                                f'("{column}" = %s AND "{related_column}" = %s)'
                                            )
                                            filter_params.append(mismatch_pair.get(column))
                                            filter_params.append(mismatch_pair.get(related_column))

                                        # Join with OR
                                        where_conditions.append(f'({" OR ".join(mismatch_conditions)})')
                            except (json.JSONDecodeError, AttributeError, KeyError):
                                # If JSON parsing fails, skip this filter
                                continue
                        else:
                            where_conditions.append(f'"{column}" {operator} %s')
                            filter_params.append(value)

                # Build ORDER BY clause
                if sort_column:
                    'ASC' if sort_direction.lower() == 'asc' else 'DESC'

                # Build the complete query
                'WHERE ' + ' AND '.join(where_conditions) if where_conditions else ''

                # Construct the table reference with schema

                query = '''
                    SELECT {', '.join([f'"{col[0]}"' for col in table_columns])}
                    FROM {table_reference}
                    {where_clause}
                    {order_by}
                '''

                # Execute query
                cursor.execute(query, filter_params)
                rows = cursor.fetchall()

                # Get headers
                headers = [col[0] for col in table_columns]

                # Convert rows to list of dictionaries for date formatting
                data = []
                for row in rows:
                    row_dict = {}
                    for i, value in enumerate(row):
                        row_dict[headers[i]] = value
                    data.append(row_dict)

                # Format date/timestamp columns based on user's preferred date format
                user_strftime_format = convert_user_date_format_to_strftime(user.date_format)
                data = format_date_columns(data, table_columns, user_strftime_format)

                # Convert back to rows format
                formatted_rows = []
                for row_dict in data:
                    formatted_rows.append([row_dict.get(header) for header in headers])

                # Generate file content based on format
                if file_format in ['csv']:
                    file_content, content_type, file_extension = self._generate_csv(headers, formatted_rows)
                elif file_format in ['excel', 'xlsx']:
                    file_content, content_type, file_extension = self._generate_excel(headers, formatted_rows)
                elif file_format == 'tsv':
                    file_content, content_type, file_extension = self._generate_tsv(headers, formatted_rows)

            customer_connection.close()

            # Create HTTP response with appropriate content
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{table_name}_data.{file_extension}"'

            return response

        except Exception as e:
            return Response(
                {"error": f"Failed to download data: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _generate_csv(self, headers, rows):
        """Generate CSV content"""
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header row
        writer.writerow(headers)

        # Write data rows
        for row in rows:
            # Convert None values to 'None' string for CSV
            csv_row = [str(val) if val is not None else 'None' for val in row]
            writer.writerow(csv_row)

        content = output.getvalue()
        output.close()

        return content, 'text/csv', 'csv'

    def _generate_tsv(self, headers, rows):
        """Generate TSV (Tab-Separated Values) content"""

        output = io.StringIO()

        # Write header row
        output.write('\t'.join(headers) + '\n')

        # Write data rows
        for row in rows:
            # Convert None values to 'None' string and join with tabs
            tsv_row = [str(val) if val is not None else 'None' for val in row]
            output.write('\t'.join(tsv_row) + '\n')

        content = output.getvalue()
        output.close()

        return content, 'text/tab-separated-values', 'tsv'

    def _generate_excel(self, headers, rows):
        """Generate Excel content"""
        try:
            from io import BytesIO

            import openpyxl
        except ImportError:
            # Fallback to basic CSV if openpyxl is not available
            return self._generate_csv(headers, rows)

        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"

        # Write headers
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Write data rows
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                # Convert None values to 'None' string for Excel
                cell_value = 'None' if value is None else value
                worksheet.cell(row=row_num, column=col_num, value=cell_value)

        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        content = output.getvalue()
        output.close()

        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx'

class LoginView(APIView):
    """
    API endpoint for user login with JWT authentication.
    """
    authentication_classes = []  # No authentication required for login
    permission_classes = []  # No permissions required for login

    def post(self, request):
        """
        Handle user login and return JWT tokens.
        """
        try:
            email = request.data.get('email')
            password = request.data.get('password')

            if not email or not password:
                return Response({
                    'error': 'Please provide both email and password'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Normalize email to lowercase and trim whitespace
            email = email.lower().strip() if email else ''
            print(f"Email received (normalized): '{email}'")

            # Validate email format
            if not email or '@' not in email:
                return Response({
                    'error': 'Invalid email format'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Custom authentication using raw SQL to avoid id column issue
            # Fetch user directly from database
            user = None
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT email, password, is_active, is_superuser, is_staff,
                               first_name, last_name, created_on, cust_id_id
                        FROM "GENERAL"."user"
                        WHERE email = %s
                    """, [email])
                    row = cursor.fetchone()

                    if not row:
                        print(f"User not found for email: {email}")
                        return Response({
                            'error': f'No account found with email: {email}. Please check your email address.'
                        }, status=status.HTTP_401_UNAUTHORIZED)

                    db_email, db_password, is_active, is_superuser, is_staff, first_name, last_name, created_on, cust_id = row

                    # Check password using User model's check_password method
                    # Create a temporary user object to use check_password
                    from api.models import User
                    # Use a temporary user ONLY for password verification
                    temp_user = User()
                    temp_user.email = db_email
                    temp_user.password = db_password
                    temp_user.created_on = created_on
                    temp_user.is_active = is_active
                    temp_user.is_superuser = is_superuser
                    temp_user.is_staff = is_staff
                    temp_user.first_name = first_name
                    temp_user.last_name = last_name

                    # Verify password
                    print(f"Checking password for user: {db_email}")
                    print(f"Password from DB length: {len(db_password) if db_password else 0}")
                    print(f"Password from DB preview: {db_password[:50] if db_password else 'None'}")
                    print(f"Created_on: {created_on}")

                    password_check_result = temp_user.check_password(password)
                    print(f"Password check result: {password_check_result}")

                    if not password_check_result:
                        print(f"Password check failed for user: {db_email}")
                        return Response({
                            'error': 'Invalid email or password'
                        }, status=status.HTTP_401_UNAUTHORIZED)

                    print(f"Password check passed for user: {db_email}")

                    # Use the SAME pattern as before: an in-memory user object,
                    # with email as the identifier for JWT (no ORM hit on legacy table).
                    user = temp_user
                    user.pk = db_email
                    user._state.adding = False
                    user._state.db = 'default'

            except Exception as auth_error:
                import traceback
                error_trace = traceback.format_exc()
                print(f"Authentication error: {auth_error}")
                print(f"Traceback: {error_trace}")
                return Response({
                    'error': f'Authentication failed: {auth_error!s}'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # Check if user was successfully authenticated
            if user is None:
                return Response({
                    'error': 'Authentication failed: User not found'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # User is authenticated, check if active
            if user.is_active:
                # Update last_login field using raw SQL (skip if fails due to schema issues)
                # We can't use user.save() because user.pk is email (string), not integer id
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            UPDATE "GENERAL"."user"
                            SET last_login = %s
                            WHERE email = %s
                        """, [timezone.now(), user.email])
                except Exception as save_error:
                    print(f"Warning: Could not update last_login: {save_error}")
                    # Continue anyway - this is not critical

                # Generate JWT tokens manually to avoid OutstandingToken issues
                # Since user table doesn't have id column, we create tokens manually
                from rest_framework_simplejwt.tokens import RefreshToken

                # Create refresh token manually
                refresh = RefreshToken()
                # Set token claims based on SIMPLE_JWT settings
                refresh['user_id'] = user.email  # USER_ID_FIELD is 'email'
                refresh['email'] = user.email
                refresh['is_superuser'] = user.is_superuser
                refresh['is_staf'] = user.is_staff

                # Access token is automatically generated from refresh token
                access_token = refresh.access_token

                getattr(user, 'cust_id', None)

                # Create response
                response = Response({
                    'message': 'Login successful',
                    'user': {
                        'email': user.email,
                        'first_name': user.first_name,
                        'last_name': user.last_name,
                    }
                }, status=status.HTTP_200_OK)

                # Set HTTP-only cookies for tokens
                response.set_cookie(
                    key='access_token',
                    value=str(access_token),
                    httponly=True,
                    secure=False,  # Set to True in production with HTTPS
                    samesite='Lax',
                    max_age=3600  # 1 hour (matches ACCESS_TOKEN_LIFETIME)
                )

                response.set_cookie(
                    key='refresh_token',
                    value=str(refresh),
                    httponly=True,
                    secure=False,  # Set to True in production with HTTPS
                    samesite='Lax',
                    max_age=86400  # 1 day (matches REFRESH_TOKEN_LIFETIME)
                )

                return response
            else:
                return Response({
                    'error': 'User account is disabled'
                }, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Login error: {e!s}")
            print(f"Traceback: {error_trace}")
            return Response({
                'error': f'Internal server error: {e!s}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class LogoutView(APIView):
    """
    API endpoint for user logout.
    Clears JWT tokens stored in cookies and Django session.
    """
    authentication_classes = []  # No authentication required for logout
    permission_classes = []  # No permissions required for logout

    def get(self, request):
        """
        Handle user logout by clearing JWT token cookies and Django session.
        """
        try:

            # Flush the Django session if it exists
            if hasattr(request, 'session'):
                request.session.flush()
                print("Django session flushed")

            # Create response
            response = Response({
                'message': 'Logout successful'
            }, status=status.HTTP_200_OK)

            # Delete the access_token cookie
            # IMPORTANT: All parameters must match those used when setting the cookie
            response.delete_cookie(
                key='access_token',
                path='/',
                samesite='Lax'
            )

            # Delete the refresh_token cookie
            # IMPORTANT: All parameters must match those used when setting the cookie
            response.delete_cookie(
                key='refresh_token',
                path='/',
                samesite='Lax'
            )

            # Also delete the sessionid cookie (Django's default session cookie)
            response.delete_cookie(
                key='sessionid',
                path='/',
                samesite='Lax'
            )

            # Delete CSRF cookie
            response.delete_cookie(
                key='csrftoken',
                path='/',
                samesite='Lax'
            )

            return response

        except Exception as e:
            return Response({
                'error': f'Logout failed: {e!s}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class RefreshTokenView(APIView):
    """
    API endpoint to refresh access token using the refresh token stored in an HttpOnly cookie.
    """
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        try:
            refresh_token = request.COOKIES.get('refresh_token')
            if not refresh_token:
                return Response({'error': 'Refresh token missing'}, status=status.HTTP_401_UNAUTHORIZED)

            serializer = TokenRefreshSerializer(data={'refresh': refresh_token})
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data

            response = Response({'message': 'Token refreshed successfully'}, status=status.HTTP_200_OK)

            # Set new access token
            response.set_cookie(
                key='access_token',
                value=data.get('access'),
                httponly=True,
                secure=False,  # Set to True in production with HTTPS
                samesite='Lax',
                max_age=3600
            )

            # If rotation is enabled, a new refresh may be returned; update cookie
            new_refresh = data.get('refresh')
            if new_refresh:
                response.set_cookie(
                    key='refresh_token',
                    value=new_refresh,
                    httponly=True,
                    secure=False,
                    samesite='Lax',
                    max_age=86400
                )

            return response
        except Exception as e:
            return Response({'error': f'Invalid or expired refresh token: {e!s}'}, status=status.HTTP_401_UNAUTHORIZED)

class TruncateTableView(APIView):
    """
    API endpoint to truncate a table.
    """
    authentication_classes = [JWTCookieAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        table_name = request.data.get('table_name')
        schema = request.data.get('schema')
        filters = request.data.get('filters', [])

        if not table_name:
            return Response(
                {"error": "table_name is required"},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            user = request.user
            if not user.cust_id:
                return Response(
                    {"error": "User is not associated with any customer."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            customer = user.cust_id
            cust_db = customer.cust_db
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND)
        except Customer.DoesNotExist:
            return Response(
                {"error": "Customer not found"},
                status=status.HTTP_404_NOT_FOUND)

        try:
            customer_connection = psycopg2.connect(
                host=settings.DATABASES['default']['HOST'],
                port=settings.DATABASES['default']['PORT'],
                database=cust_db,
                user=settings.DATABASES['default']['USER'],
                password=settings.DATABASES['default']['PASSWORD']
            )
            customer_connection.autocommit = True

            with customer_connection.cursor() as cursor:
                # Construct the table reference with schema
                table_reference = f'"{schema}"."{table_name}"'

                # Get count before truncation/deletion
                cursor.execute(f'SELECT COUNT(*) FROM {table_reference}')
                records_count = cursor.fetchone()[0]

                if filters:
                    where_conditions = []
                    filter_params = []
                    for filter in filters:
                        column = filter.get('column')
                        operator = filter.get('operator')
                        value = filter.get('value')

                        if operator == 'ILIKE':
                            where_conditions.append(f'"{column}" ILIKE %s')
                            filter_params.append(f'%{value}%')
                        elif operator == 'IN':
                            value_list = [v.strip() for v in value.split(',') if v.strip()]
                            placeholders = ','.join(['%s'] * len(value_list))
                            where_conditions.append(f'"{column}" IN ({placeholders})')
                            filter_params.extend(value_list)
                        elif operator == 'NOT IN':
                            value_list = [v.strip() for v in value.split(',') if v.strip()]
                            placeholders = ','.join(['%s'] * len(value_list))
                            where_conditions.append(f'"{column}" NOT IN ({placeholders})')
                            filter_params.extend(value_list)
                        elif operator == 'IS NULL':
                            where_conditions.append(f'"{column}" IS NULL')
                        elif operator == 'IS NOT NULL':
                            where_conditions.append(f'"{column}" IS NOT NULL')

                        else:
                            where_conditions.append(f'"{column}" {operator} %s')
                            filter_params.append(value)

                    cursor.execute(f'SELECT COUNT(*) FROM {table_reference} WHERE {" AND ".join(where_conditions)}', filter_params)
                    deleted_count = cursor.fetchone()[0]
                    delete_query = f'DELETE FROM {table_reference} WHERE {" AND ".join(where_conditions)}'
                    cursor.execute(delete_query, filter_params)

                else:
                    cursor.execute(f'TRUNCATE TABLE {table_reference}')
                    deleted_count = records_count

                # Log the truncate/delete activity
                try:
                    insert_activity_log_sql = '''
                        INSERT INTO "GENERAL"."activity_log" (table_name, created_by, records_count, modified_by, description)
                        VALUES (%s, %s, %s, %s, %s)
                    '''
                    if filters:
                        description = f"Deleted {deleted_count} record(s) from table '{table_name}' with filters"
                    else:
                        description = f"Table '{table_name}' truncated with {deleted_count} record(s) removed"

                    cursor.execute(
                        insert_activity_log_sql,
                        (table_name, None, deleted_count, user.email, description)
                    )
                except Exception as log_error:
                    print(f"Error logging truncate activity: {log_error}")

            customer_connection.close()

            return Response({"message": "Table truncated successfully"}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"Failed to truncate table: {e!s}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR)
